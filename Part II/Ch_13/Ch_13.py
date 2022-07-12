"""
WORKING WITH EXCEL SPREADSHEETS

Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 301). No Starch Press. Kindle Edition. 
"""

import openpyxl

# Opening Excel Documents with OpenPyXL

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 303). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
type(wb)

# Getting Sheets from the Workbook

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 304). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames   # the workbook's sheets' names
sheet = wb['Sheet3']    # get a sheet from the workbook
sheet
type(sheet)
sheet.title # Get the sheet's title as a string
anotherSheet = wb.active    # Get the active sheet
anotherSheet

# Getting Cells from the Sheets

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 304). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']    # Get the sheet
sheet['A1'] # Get the cell from the sheet
sheet['A1'].value   # Get the value
c = sheet['B1'] # Get another cell from the sheet
c.value
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
'Cell %s is %s' % (c.coordinate, c.value)
sheet['C1'].value

sheet.cell(row=1, column=2)
sheet.cell(row=1, column=2).value
for i in range(1, 8, 2):    # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
sheet.max_row   # Get the highest row number
sheet.max_column    # Get the highest column number NOTE: it is a number, not a letter

# Converting Between Column Letters and Numbers

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 306). No Starch Press. Kindle Edition. 

import openpyxl

from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(2)
get_column_letter(27)
get_column_letter(1900)
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
column_index_from_string('A')   # Get A's number
column_index_from_string('ZZZ')

# Getting Rows and Columns from the Sheets

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 306). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
    
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
list(sheet.columns)[1]
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
    
# Creating and Saving Excel Documents

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 313). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.Workbook()    # Create a blank workbook
wb.sheetnames   # It starts with one sheet.
sheet = wb.active
sheet.title
sheet.title = 'Spam Bacon Eggs Sheet'   # Change title
wb.sheetnames

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Soam Spam'
wb.save('example_copy.xlsx')    # Save the workbook

# Creating and Removing Sheets

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 314). No Starch Press. Kindle Edition. 

import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
wb.create_sheet()   # Add a new sheet
wb.sheetnames
# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')
wb.sheetnames
wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames

wb.sheetnames
del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames

wb.save('example_copy.xlsx')

# Writing Values to Cells

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 314). No Starch Press. Kindle Edition. 

import openpyxl

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'   # Edit the cell's value
sheet['A1'].value

# Setting the Font Style of Cells

# Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 318). No Starch Press. Kindle Edition. 

from openpyxl.styles import Font

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24font = Font(size=24, italic=True)   # Create a font
sheet['A1'].font = italic24font # Apply the font to A1
sheet['A1'] = 'Hello, world!'
wb.save('styles.xlsx')