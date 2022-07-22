"""
sendDuesReminders.py

Sweigart, Al. Automate the Boring Stuff with Python, 2nd Edition (p. 434). No Starch Press. Kindle Edition. 
"""

#!
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each memeber's payment status

unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to example email account

smtpObj = smtplib.SMTP('smtp.example.comm', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('none@none.com', sys.argv[1])

# Send out remidner emails 

for name, email in unpaidMembers.items():
    body = "Subject %s dues unpaid.\nPay the bill. thx" % (latestMonth, name, latestMonth)
    print("sending email to %s..." % email)
    sendmailStatus = smtpObj.sendmail('none@none.com', email, body)
    if sendmailStatus != {}:
        print('problem sending email to %s: %s' % (email, sendmailStatus))
        
smtpObj.quit()
    